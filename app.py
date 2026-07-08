"""
app.py
------
Dashboard de diagnostic automatisé des temps d'arrêt (downtime)
Projet PFA - Génie Mécatronique - Cas Versigent

Lancer avec :
    streamlit run app.py
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
from openpyxl.utils import get_column_letter

from utils import (
    load_data, standardize_columns, clean_data,
    compute_pareto, compute_pareto_level2, compute_paynter,
    compute_frequency_pareto,
    compute_alerts, predict_next_week, summary_kpis,
    generate_text_summary, build_action_plan_template,
    build_action_plan_par_famille,
    format_excel_sheet, REQUIRED_COLUMNS,
    auto_detect_family_column,
    add_seuil80_column, add_pareto_excel_chart, add_repartition_chart_excel,
    build_pptx_report,
)

st.set_page_config(page_title="Dashboard Downtime - Versigent", layout="wide", page_icon="🛠️")


def fig_to_png_bytes(fig):
    """Convertit un graphique Plotly en image PNG téléchargeable (pour rapport/PPT)."""
    try:
        return fig.to_image(format="png", scale=2, width=1100, height=550)
    except Exception:
        return None

# ---------------------------------------------------------------------------
# EN-TÊTE
# ---------------------------------------------------------------------------
st.title("🛠️ Dashboard de diagnostic automatisé des temps d'arrêt")
st.caption(
    "Projet PFA — Génie Mécatronique | Remplace le calcul manuel Pareto / 5 Why / "
    "Paynter Chart réalisé aujourd'hui dans Excel."
)

with st.expander("ℹ️ Comment utiliser ce dashboard (à lire la première fois)"):
    st.markdown("""
    1. Charge un fichier Excel/CSV de pannes dans la barre latérale (ou utilise le
       **jeu de données d'exemple** pour tester tout de suite).
    2. Si les noms de colonnes de ton fichier sont différents, fais la correspondance
       dans la section **"Correspondance des colonnes"**.
    3. Les graphiques (Pareto, Paynter) et les alertes se génèrent automatiquement.
    4. Tu peux exporter les tableaux calculés en Excel via les boutons de téléchargement.
    """)

# ---------------------------------------------------------------------------
# BARRE LATÉRALE : CHARGEMENT DES DONNÉES
# ---------------------------------------------------------------------------
st.sidebar.header("1. Données")

use_sample = st.sidebar.checkbox("Utiliser le jeu de données d'exemple (Versigent)", value=False)

uploaded_file = None
if not use_sample:
    uploaded_file = st.sidebar.file_uploader(
        "Charger un fichier de pannes (.xlsx, .xls ou .csv)", type=["xlsx", "xls", "csv"]
    )

raw_df = None
if use_sample:
    try:
        raw_df = pd.read_excel("donnees_pannes_exemple.xlsx")
        st.sidebar.success("Jeu de données d'exemple chargé (basé sur les causes réelles Versigent).")
    except FileNotFoundError:
        st.sidebar.error(
            "Le fichier d'exemple « donnees_pannes_exemple.xlsx » est introuvable sur le serveur. "
            "Décoche la case ci-dessus et charge ton propre fichier, ou ajoute "
            "« donnees_pannes_exemple.xlsx » dans le même dossier GitHub que app.py."
        )
elif uploaded_file is not None:
    try:
        raw_df = load_data(uploaded_file)
        st.sidebar.success(f"Fichier chargé : {len(raw_df)} lignes.")
    except Exception as e:
        st.sidebar.error(f"Erreur de lecture du fichier : {e}")

if raw_df is None:
    st.info("👈 Charge un fichier de données ou coche la case d'exemple pour démarrer.")
    st.stop()

# ---------------------------------------------------------------------------
# CORRESPONDANCE DES COLONNES (rend l'outil compatible avec n'importe quel export Versigent)
# ---------------------------------------------------------------------------
st.sidebar.header("2. Correspondance des colonnes")
cols = list(raw_df.columns)

def guess(col_name, candidates):
    """Devine la colonne correspondant à `candidates` (mots-clés, du plus prioritaire
    au moins prioritaire). Retourne None si rien ne correspond, pour éviter de
    retomber silencieusement sur une colonne au hasard (ex: 'Site Name')."""
    for c in candidates:
        for col in cols:
            if c.lower() in str(col).lower():
                return col
    return None

def guessed_index(guess_result):
    """Index à utiliser dans le selectbox : celui de la colonne devinée si trouvée,
    sinon 0 par défaut (l'utilisateur devra vérifier/corriger manuellement)."""
    return cols.index(guess_result) if guess_result in cols else 0

mapping_ui = {}
mapping_ui["Date"] = st.sidebar.selectbox(
    "Colonne Date", cols, index=guessed_index(guess("Date", ["date"]))
)
mapping_ui["Machine"] = st.sidebar.selectbox(
    "Colonne Machine", cols,
    index=guessed_index(guess("Machine", [
        "asset description", "machine", "equipement", "équipement", "asset"
    ])),
    help="Choisis la colonne qui identifie l'équipement précis en panne "
         "(ex: 'Asset Description' : M02, V827G, Kit Seal 24646...). "
         "Évite 'Asset ID' ou 'Asset / Position', qui sont des codes internes "
         "moins lisibles dans les graphiques et tableaux."
)
mapping_ui["Cause"] = st.sidebar.selectbox(
    "Colonne Cause", cols,
    index=guessed_index(guess("Cause", ["fault code", "code défaut", "cause", "reason", "défaut", "defaut"])),
    help="Choisis la colonne qui contient le CODE ou la CATÉGORIE de la panne "
         "(ex: 'Fault Code' : CRMP-Crimping Problem, NOGO-Will Not Start...). "
         "Ne choisis pas une colonne comme 'Site Name' ou 'Job ID' qui n'a rien à voir "
         "avec la cause de la panne."
)
mapping_ui["Duree_min"] = st.sidebar.selectbox(
    "Colonne Durée (minutes)", cols,
    index=guessed_index(guess("Duree", ["total duration", "duree", "durée", "duration"]))
)

# Colonne "Famille" (catégorie de machine : Cutting Machine, Kit Seal, Outils, Press...)
# Optionnelle : si le fichier ne l'a pas, le dashboard fonctionne comme avant (Pareto global).
# Si elle est présente, on active en plus l'analyse par famille demandée par l'entreprise
# (ne jamais mélanger toutes les pannes dans un seul Pareto).
family_guess = auto_detect_family_column(cols)
use_family = st.sidebar.checkbox(
    "Mon fichier a une colonne Famille / Catégorie de machine (ex: Sub Description)",
    value=family_guess is not None,
    help="Active le Pareto par famille (Cutting Machine, Kit Seal, Outils, Press...) demandé par Versigent, "
         "au lieu d'un seul Pareto qui mélange toutes les pannes."
)
if use_family:
    mapping_ui["Famille"] = st.sidebar.selectbox(
        "Colonne Famille / Catégorie", cols, index=guessed_index(family_guess)
    )

# Inverse le mapping : {nom_colonne_fichier: nom_standard}
inverse_mapping = {v: k for k, v in mapping_ui.items()}

try:
    df = standardize_columns(raw_df, inverse_mapping)
    df = clean_data(df)
except Exception as e:
    st.error(f"Erreur de préparation des données : {e}")
    st.stop()

if df.attrs.get("lignes_supprimees", 0) > 0:
    st.sidebar.warning(f"{df.attrs['lignes_supprimees']} ligne(s) invalide(s) ignorée(s) (date ou durée manquante).")

if df.attrs.get("dates_corrigees", 0) > 0:
    st.sidebar.warning(
        f"🗓️ {df.attrs['dates_corrigees']} date(s) avaient le jour et le mois inversés "
        f"(bug d'export courant) — corrigées automatiquement pour que le Paynter chart et "
        f"les filtres par période restent fiables."
    )

# ---------------------------------------------------------------------------
# GARDE-FOU : détecte une mauvaise correspondance de colonne, quel que soit le
# fichier chargé. Si une colonne clé (Cause, Machine, Famille) n'a qu'une seule
# valeur unique, le Pareto correspondant n'aura aucun sens (ex: mapper "Cause"
# sur "Site Name" qui vaut toujours "Morocco V") : on prévient clairement au
# lieu de laisser un résultat trompeur passer inaperçu.
# ---------------------------------------------------------------------------
_colonnes_a_verifier = [("Cause", "Cause"), ("Machine", "Machine")]
if "Famille" in df.columns:
    _colonnes_a_verifier.append(("Famille", "Famille / Catégorie"))

for col_standard, label in _colonnes_a_verifier:
    if col_standard in df.columns and df[col_standard].nunique(dropna=True) <= 1:
        valeur_unique = df[col_standard].iloc[0] if not df.empty else "?"
        st.warning(
            f"⚠️ La colonne mappée sur **{label}** (« {mapping_ui.get(col_standard, '?')} ») "
            f"n'a qu'une seule valeur pour toutes les pannes (« {valeur_unique} »). "
            f"Le Pareto correspondant n'aura pas de sens. Vérifie la correspondance des colonnes "
            f"dans la barre latérale (section 2) — choisis une colonne comme 'Fault Code' pour la Cause, "
            f"ou 'Asset Description' pour la Machine."
        )

# ---------------------------------------------------------------------------
# FILTRES
# ---------------------------------------------------------------------------
st.sidebar.header("3. Filtres")
machines = sorted(df["Machine"].unique().tolist())
selected_machines = st.sidebar.multiselect("Filtrer par machine", machines, default=machines)
df = df[df["Machine"].isin(selected_machines)]

date_min, date_max = df["Date"].min(), df["Date"].max()
if pd.notna(date_min) and pd.notna(date_max):
    date_range = st.sidebar.date_input("Période", value=(date_min.date(), date_max.date()))
    if len(date_range) == 2:
        df = df[(df["Date"] >= pd.to_datetime(date_range[0])) & (df["Date"] <= pd.to_datetime(date_range[1]))]

if df.empty:
    st.warning("Aucune donnée pour ces filtres.")
    st.stop()

# ---------------------------------------------------------------------------
# KPIs
# ---------------------------------------------------------------------------
kpis = summary_kpis(df)
c1, c2, c3, c4 = st.columns(4)
c1.metric("Nombre de pannes", kpis["nb_evenements"])
c2.metric("Downtime total", f"{kpis['duree_totale_h']} h")
c3.metric("Machine la plus critique", kpis["machine_top"])
c4.metric("Cause la plus fréquente", kpis["cause_top"])

pareto1_preview = compute_pareto(df, group_col="Cause")
resume_text = generate_text_summary(df, pareto1_preview, group_col="Cause")
st.info(f"📝 **Résumé automatique** (copiable pour ton rapport) :\n\n{resume_text}")

st.divider()

# ---------------------------------------------------------------------------
# PARETO NIVEAU 1
# ---------------------------------------------------------------------------
st.header("📊 Pareto Niveau 1 — Causes de panne")

pareto1 = pareto1_preview

fig1 = go.Figure()
fig1.add_bar(x=pareto1["Cause"], y=pareto1["Duree_totale_min"], name="Durée (min)", marker_color="#1f77b4")
fig1.add_trace(go.Scatter(
    x=pareto1["Cause"], y=pareto1["Cumul_%"], name="Cumul %",
    yaxis="y2", mode="lines+markers", line=dict(color="red")
))
fig1.add_hline(y=80, line_dash="dash", line_color="gray", yref="y2")
fig1.update_layout(
    yaxis=dict(title="Durée totale (min)"),
    yaxis2=dict(title="Cumul %", overlaying="y", side="right", range=[0, 100]),
    xaxis=dict(tickangle=-45),
    legend=dict(orientation="h", y=1.1),
    height=450,
)
st.plotly_chart(fig1, use_container_width=True)

with st.expander("Voir le tableau Pareto niveau 1"):
    st.dataframe(pareto1, use_container_width=True)

# Export
col_exp1, col_exp2 = st.columns(2)
buf1 = BytesIO()
pareto1.to_excel(buf1, index=False)
col_exp1.download_button("⬇️ Télécharger le tableau (Excel)", buf1.getvalue(),
                          file_name="pareto_niveau1.xlsx", key="dl_pareto1_xlsx")

png1 = fig_to_png_bytes(fig1)
if png1:
    col_exp2.download_button("🖼️ Télécharger le graphique (image PNG)", png1,
                              file_name="pareto_niveau1.png", mime="image/png", key="dl_pareto1_png")

# ---------------------------------------------------------------------------
# PARETO NIVEAU 2
# ---------------------------------------------------------------------------
st.header("🔍 Pareto Niveau 2 — Détail par machine pour une cause")

top_cause = st.selectbox("Choisir la cause à décomposer", pareto1["Cause"].tolist())
pareto2 = compute_pareto_level2(df, top_value=top_cause, level1_col="Cause", level2_col="Machine")

if not pareto2.empty:
    # Même style de vrai Pareto que le Niveau 1 (barres + courbe de cumul % + seuil 80%),
    # au lieu d'un simple bar chart qui ne permettait pas de voir la règle 80/20 à ce niveau.
    fig2 = go.Figure()
    fig2.add_bar(x=pareto2["Machine"], y=pareto2["Duree_totale_min"], name="Durée (min)", marker_color="#1f77b4")
    fig2.add_trace(go.Scatter(
        x=pareto2["Machine"], y=pareto2["Cumul_%"], name="Cumul %",
        yaxis="y2", mode="lines+markers", line=dict(color="red")
    ))
    fig2.add_hline(y=80, line_dash="dash", line_color="gray", yref="y2")
    fig2.update_layout(
        title=f"Pareto — Détail machine pour la cause : {top_cause}",
        yaxis=dict(title="Durée totale (min)"),
        yaxis2=dict(title="Cumul %", overlaying="y", side="right", range=[0, 100]),
        xaxis=dict(tickangle=-45),
        legend=dict(orientation="h", y=1.1),
        height=450,
    )
    st.plotly_chart(fig2, use_container_width=True)
    st.dataframe(pareto2, use_container_width=True)

    col_exp3, col_exp4 = st.columns(2)
    buf2 = BytesIO()
    pareto2.to_excel(buf2, index=False)
    col_exp3.download_button("⬇️ Télécharger le tableau (Excel)", buf2.getvalue(),
                              file_name="pareto_niveau2.xlsx", key="dl_pareto2_xlsx")

    png2 = fig_to_png_bytes(fig2)
    if png2:
        col_exp4.download_button("🖼️ Télécharger le graphique (image PNG)", png2,
                                  file_name="pareto_niveau2.png", mime="image/png", key="dl_pareto2_png")
else:
    st.info("Pas assez de données pour cette cause.")

st.divider()

# ---------------------------------------------------------------------------
# PARETO PAR FAMILLE (format demandé par l'entreprise : ne pas englober toutes
# les pannes dans un seul Pareto, mais faire un Pareto par famille de machine :
# Cutting Machine, Kit Seal, Outils, Press... comme dans "Cutting Analysis Week 27")
# ---------------------------------------------------------------------------
st.header("🏭 Pareto par Famille de machines (format Versigent « Cutting Analysis »)")
st.caption(
    "Au lieu d'un seul Pareto qui mélange toutes les pannes, chaque famille de machine "
    "(Cutting Machine, Kit Seal, Outils, Press...) a son propre Pareto et son propre "
    "camembert de répartition — exactement comme dans le rapport Cutting Analysis."
)

pareto_famille = pd.DataFrame()
pareto_par_famille_detail = {}
pareto_par_famille_figs = {}
fig_fam1 = None
fig_fam_pie = None

if "Famille" not in df.columns:
    st.info(
        "Aucune colonne Famille sélectionnée. Coche la case « Mon fichier a une colonne "
        "Famille / Catégorie de machine » dans la barre latérale (section 2) pour activer "
        "cette analyse."
    )
else:
    # --- Pareto niveau 1 : par famille ---
    pareto_famille = compute_pareto(df, group_col="Famille")

    fig_fam1 = go.Figure()
    fig_fam1.add_bar(x=pareto_famille["Famille"], y=pareto_famille["Duree_totale_min"],
                      name="Durée (min)", marker_color="#1f77b4")
    fig_fam1.add_trace(go.Scatter(
        x=pareto_famille["Famille"], y=pareto_famille["Cumul_%"], name="Cumul %",
        yaxis="y2", mode="lines+markers", line=dict(color="red")
    ))
    fig_fam1.add_hline(y=80, line_dash="dash", line_color="gray", yref="y2")
    fig_fam1.update_layout(
        title="Pareto Niveau 1 — Répartition du downtime par famille",
        yaxis=dict(title="Durée totale (min)"),
        yaxis2=dict(title="Cumul %", overlaying="y", side="right", range=[0, 100]),
        xaxis=dict(tickangle=-45),
        legend=dict(orientation="h", y=1.1),
        height=450,
    )

    fig_fam_pie = px.pie(
        pareto_famille, names="Famille", values="Duree_totale_min", hole=0.35,
        title="Répartition du downtime par famille (%)"
    )
    fig_fam_pie.update_traces(textposition="inside", textinfo="percent+label")

    col_fam1, col_fam2 = st.columns(2)
    col_fam1.plotly_chart(fig_fam1, use_container_width=True)
    col_fam2.plotly_chart(fig_fam_pie, use_container_width=True)

    with st.expander("Voir le tableau Pareto par famille"):
        st.dataframe(pareto_famille, use_container_width=True)

    st.subheader("🔍 Pareto Niveau 2 — Détail par machine, pour chaque famille")

    familles_list = pareto_famille["Famille"].tolist()
    tabs = st.tabs(familles_list)
    pareto_par_famille_detail = {}
    pareto_par_famille_figs = {}  # conservés pour être réutilisés dans le PPT plus bas

    for tab, famille_name in zip(tabs, familles_list):
        with tab:
            sous_pareto = compute_pareto_level2(
                df, top_value=famille_name, level1_col="Famille", level2_col="Machine"
            )
            if sous_pareto.empty:
                st.info(f"Pas assez de données pour la famille « {famille_name} ».")
                continue

            pareto_par_famille_detail[famille_name] = sous_pareto

            fig_sub_bar = go.Figure()
            fig_sub_bar.add_bar(x=sous_pareto["Machine"], y=sous_pareto["Duree_totale_min"],
                                 name="Durée (min)", marker_color="#1f77b4")
            fig_sub_bar.add_trace(go.Scatter(
                x=sous_pareto["Machine"], y=sous_pareto["Cumul_%"], name="Cumul %",
                yaxis="y2", mode="lines+markers", line=dict(color="red")
            ))
            fig_sub_bar.add_hline(y=80, line_dash="dash", line_color="gray", yref="y2")
            fig_sub_bar.update_layout(
                title=f"Pareto — {famille_name}",
                yaxis=dict(title="Durée totale (min)"),
                yaxis2=dict(title="Cumul %", overlaying="y", side="right", range=[0, 100]),
                xaxis=dict(tickangle=-45),
                legend=dict(orientation="h", y=1.1),
                height=420,
            )

            fig_sub_pie = px.pie(
                sous_pareto, names="Machine", values="Duree_totale_min", hole=0.35,
                title=f"Répartition (%) — {famille_name}"
            )
            fig_sub_pie.update_traces(textposition="inside", textinfo="percent+label")

            pareto_par_famille_figs[famille_name] = (fig_sub_bar, fig_sub_pie)

            col_sub1, col_sub2 = st.columns(2)
            col_sub1.plotly_chart(fig_sub_bar, use_container_width=True)
            col_sub2.plotly_chart(fig_sub_pie, use_container_width=True)

            st.dataframe(sous_pareto, use_container_width=True)

    # --- Export Excel complet : synthèse + une feuille par famille, avec de VRAIS
    # graphiques Excel natifs (barres + courbe de cumul % + seuil 80%, comme dans
    # les sections "TAB"/"Cutting Analysis" du fichier de référence de l'entreprise)
    # au lieu de simples tableaux — c'est ce qui rendait le fichier exporté différent
    # du format attendu jusqu'ici.
    freq_pareto = compute_frequency_pareto(df, group_col="Machine")

    buf_famille = BytesIO()
    with pd.ExcelWriter(buf_famille, engine="openpyxl") as writer:
        # Synthèse Familles : tableau + Pareto niveau 1 (par famille) + donut de répartition
        pareto_famille.to_excel(writer, sheet_name="Synthèse Familles", index=False)
        ws_synth = writer.sheets["Synthèse Familles"]
        format_excel_sheet(ws_synth, n_cols=len(pareto_famille.columns))
        seuil_synth = add_seuil80_column(ws_synth, n_rows=len(pareto_famille),
                                          table_col_count=len(pareto_famille.columns))
        add_pareto_excel_chart(
            ws_synth, header_row=1, n_rows=len(pareto_famille), cat_col=1, val_col=2,
            cumul_col=4, seuil_col=seuil_synth,
            title="Pareto Niveau 1 — Durée totale par famille",
            anchor=f"{get_column_letter(seuil_synth + 2)}1",
        )
        add_repartition_chart_excel(
            ws_synth, header_row=1, n_rows=len(pareto_famille), cat_col=1, val_col=2,
            title="Répartition du downtime par famille",
            anchor=f"{get_column_letter(seuil_synth + 2)}20",
        )

        # Une feuille détaillée par famille : tableau + Pareto niveau 2 (par machine)
        for famille_name, sous_pareto in pareto_par_famille_detail.items():
            # Les noms de feuille Excel sont limités à 31 caractères
            sheet_name = str(famille_name)[:31] if str(famille_name).strip() else "Famille"
            sous_pareto.to_excel(writer, sheet_name=sheet_name, index=False)
            ws_fam = writer.sheets[sheet_name]
            format_excel_sheet(ws_fam, n_cols=len(sous_pareto.columns))
            seuil_fam = add_seuil80_column(ws_fam, n_rows=len(sous_pareto),
                                            table_col_count=len(sous_pareto.columns))
            add_pareto_excel_chart(
                ws_fam, header_row=1, n_rows=len(sous_pareto), cat_col=1, val_col=2,
                cumul_col=4, seuil_col=seuil_fam, title=f"Pareto — {famille_name}",
                anchor=f"{get_column_letter(seuil_fam + 2)}1",
            )

        # Fréquence des pannes (toutes familles confondues), comme le tableau
        # "Frequent breakdowns" du fichier de référence
        freq_pareto.to_excel(writer, sheet_name="Fréquence pannes", index=False)
        ws_freq = writer.sheets["Fréquence pannes"]
        format_excel_sheet(ws_freq, n_cols=len(freq_pareto.columns))
        seuil_freq = add_seuil80_column(ws_freq, n_rows=len(freq_pareto),
                                         table_col_count=len(freq_pareto.columns))
        add_pareto_excel_chart(
            ws_freq, header_row=1, n_rows=len(freq_pareto), cat_col=1, val_col=2,
            cumul_col=4, seuil_col=seuil_freq, title="Fréquence des pannes par équipement",
            anchor=f"{get_column_letter(seuil_freq + 2)}1", value_axis_title="Nombre de pannes",
            width=20, height=10,
        )

        # Plan d'action pré-rempli, construit par famille (équipements les plus
        # critiques de chaque famille, comme la feuille "ACTION PLAN" de référence)
        action_plan_famille = build_action_plan_par_famille(pareto_par_famille_detail, top_n=3)
        if not action_plan_famille.empty:
            action_plan_famille.to_excel(writer, sheet_name="Plan d'action", index=False)
            format_excel_sheet(writer.sheets["Plan d'action"],
                                n_cols=len(action_plan_famille.columns), header_color="C0392B")

    st.download_button(
        "📦 Télécharger l'analyse par famille (Excel, format Cutting Analysis — avec graphiques natifs)",
        buf_famille.getvalue(),
        file_name=f"Cutting_Analysis_par_famille_{pd.Timestamp.today().date()}.xlsx",
        key="dl_famille_xlsx",
    )

st.divider()

# ---------------------------------------------------------------------------
# PARETO PAR FRÉQUENCE DES PANNES (« Frequent breakdowns » dans le fichier de
# référence de l'entreprise) : classement par NOMBRE de pannes, toutes familles
# confondues — complémentaire du Pareto par durée. Un équipement qui tombe en
# panne très souvent doit remonter ici même si chaque panne est courte.
# ---------------------------------------------------------------------------
st.header("🔁 Pareto par Fréquence des pannes — équipements qui tombent en panne le plus souvent")
st.caption(
    "Ce Pareto classe les équipements par NOMBRE de pannes (et non par durée cumulée), "
    "comme le tableau « Frequent breakdowns » du format Versigent. Un équipement peut "
    "apparaître ici même si chaque panne est courte, dès qu'il tombe en panne très souvent — "
    "ce qui révèle un problème de fiabilité récurrent, différent d'un simple gros arrêt ponctuel."
)

freq_pareto = compute_frequency_pareto(df, group_col="Machine")

fig_freq = go.Figure()
fig_freq.add_bar(x=freq_pareto["Machine"], y=freq_pareto["Nombre_pannes"],
                  name="Nombre de pannes", marker_color="#2ca02c")
fig_freq.add_trace(go.Scatter(
    x=freq_pareto["Machine"], y=freq_pareto["Cumul_%"], name="Cumul %",
    yaxis="y2", mode="lines+markers", line=dict(color="red")
))
fig_freq.add_hline(y=80, line_dash="dash", line_color="gray", yref="y2")
fig_freq.update_layout(
    title="Fréquence des pannes par équipement",
    yaxis=dict(title="Nombre de pannes"),
    yaxis2=dict(title="Cumul %", overlaying="y", side="right", range=[0, 100]),
    xaxis=dict(tickangle=-45),
    legend=dict(orientation="h", y=1.1),
    height=450,
)
st.plotly_chart(fig_freq, use_container_width=True)

with st.expander("Voir le tableau Fréquence des pannes"):
    st.dataframe(freq_pareto, use_container_width=True)

col_freq1, col_freq2 = st.columns(2)
buf_freq = BytesIO()
freq_pareto.to_excel(buf_freq, index=False)
col_freq1.download_button("⬇️ Télécharger le tableau (Excel)", buf_freq.getvalue(),
                           file_name="frequence_pannes.xlsx", key="dl_freq_xlsx")

png_freq = fig_to_png_bytes(fig_freq)
if png_freq:
    col_freq2.download_button("🖼️ Télécharger le graphique (image PNG)", png_freq,
                               file_name="frequence_pannes.png", mime="image/png", key="dl_freq_png")

st.divider()

# ---------------------------------------------------------------------------
# PAYNTER CHART
# ---------------------------------------------------------------------------
st.header("📈 Paynter Chart — Évolution hebdomadaire des causes")

paynter = compute_paynter(df, group_col="Cause")

if not paynter.empty and paynter.shape[1] > 0:
    fig3 = go.Figure()
    for cause in paynter.index:
        fig3.add_trace(go.Scatter(
            x=paynter.columns, y=paynter.loc[cause], mode="lines+markers", name=cause
        ))
    fig3.update_layout(xaxis_title="Semaine", yaxis_title="Durée (min)", height=450)
    st.plotly_chart(fig3, use_container_width=True)

    with st.expander("Voir le tableau Paynter"):
        st.dataframe(paynter, use_container_width=True)

    col_exp5, col_exp6 = st.columns(2)
    buf3 = BytesIO()
    paynter.to_excel(buf3)
    col_exp5.download_button("⬇️ Télécharger le tableau (Excel)", buf3.getvalue(),
                              file_name="paynter_chart.xlsx", key="dl_paynter_xlsx")

    png3 = fig_to_png_bytes(fig3)
    if png3:
        col_exp6.download_button("🖼️ Télécharger le graphique (image PNG)", png3,
                                  file_name="paynter_chart.png", mime="image/png", key="dl_paynter_png")
else:
    st.info("Pas assez de semaines différentes dans les données pour tracer un Paynter chart.")

st.divider()

# ---------------------------------------------------------------------------
# ALERTES
# ---------------------------------------------------------------------------
st.header("🚨 Alertes machines critiques")

threshold = st.slider("Seuil d'alerte (minutes cumulées sur la période)", min_value=10, max_value=500, value=100, step=10)
alerts = compute_alerts(df, threshold_min=threshold, group_col="Machine")

if not alerts.empty:
    st.error(f"{len(alerts)} machine(s) dépassent le seuil de {threshold} min :")
    st.dataframe(alerts, use_container_width=True)
else:
    st.success("Aucune machine ne dépasse le seuil actuel.")

st.divider()

# ---------------------------------------------------------------------------
# PRÉDICTION SIMPLE
# ---------------------------------------------------------------------------
st.header("🔮 Prédiction — downtime probable la semaine prochaine")

if not paynter.empty:
    cause_pred = st.selectbox("Choisir la cause à prédire", paynter.index.tolist(), key="pred")
    result = predict_next_week(paynter, cause_pred)
    if result["prediction"] is not None:
        st.metric(f"Prédiction pour « {cause_pred} »", f"{result['prediction']} min",
                   help=f"Méthode utilisée : {result['methode']}")
    else:
        st.info("Pas assez de données pour prédire.")
else:
    st.info("Pas assez de données pour une prédiction.")

st.divider()

# ---------------------------------------------------------------------------
# EXPORT COMPLET — PRÉSENTATION POWERPOINT (le livrable final pour Versigent)
# Remplace l'ancien rapport Excel : une présentation prête à projeter en réunion,
# avec TOUS les graphiques (mêmes couleurs que le dashboard) et tableaux calculés
# plus haut. L'analyse détaillée "format Cutting Analysis" avec graphiques Excel
# natifs reste disponible plus haut, dans la section Pareto par Famille, pour
# l'équipe qui a besoin de retravailler les chiffres directement dans Excel.
# ---------------------------------------------------------------------------
st.header("📦 Rapport PowerPoint complet — prêt à projeter en réunion")
st.caption(
    "Une seule présentation PowerPoint, prête à partager : synthèse, Pareto niveau 1 & 2, "
    "Pareto par famille, Fréquence des pannes, Paynter chart, alertes, et un plan d'action "
    "pré-rempli avec les causes/équipements prioritaires — mêmes graphiques et mêmes couleurs "
    "que ce dashboard. L'équipe maintenance n'a plus qu'à compléter les colonnes d'actions "
    "correctives directement dans les slides."
)

# Choix du plan d'action : par famille si dispo (ce que l'entreprise attend), sinon global
if pareto_par_famille_detail:
    action_plan = build_action_plan_par_famille(pareto_par_famille_detail, top_n=3)
    action_plan_ppt = action_plan[[c for c in
        ["Item Nr", "Famille", "Équipement / Machine", "Durée totale (min)", "Statut (à compléter)"]
        if c in action_plan.columns]]
    action_plan_header_color = "C0392B"
else:
    action_plan = build_action_plan_template(pareto1, group_col="Cause", top_n=5)
    action_plan_ppt = action_plan[[c for c in
        ["Priorité", "Primary Reason Code", "Root Cause Effect (%)", "Statut (à compléter)"]
        if c in action_plan.columns]]
    action_plan_header_color = "C0392B"

pptx_sections = [
    {
        "title": "Synthèse",
        "kpis": [
            ("Nombre de pannes", kpis["nb_evenements"]),
            ("Downtime total", f"{kpis['duree_totale_h']} h"),
            ("Machine la plus critique", kpis["machine_top"]),
            ("Cause la plus fréquente", kpis["cause_top"]),
        ],
        "body_text": resume_text,
    },
    {
        "title": "📊 Pareto Niveau 1 — Causes de panne",
        "image": fig_to_png_bytes(fig1),
        "table": pareto1,
    },
]

if not pareto2.empty:
    pptx_sections.append({
        "title": f"🔍 Pareto Niveau 2 — Détail machine pour la cause : {top_cause}",
        "image": fig_to_png_bytes(fig2),
        "table": pareto2,
    })

if not pareto_famille.empty:
    pptx_sections.append({
        "title": "🏭 Pareto par Famille de machines (format Versigent « Cutting Analysis »)",
        "caption": "Chaque famille de machine a son propre Pareto et son propre camembert "
                   "de répartition — au lieu d'un seul Pareto qui mélange toutes les pannes.",
        "image": fig_to_png_bytes(fig_fam1),
        "image2": fig_to_png_bytes(fig_fam_pie),
    })
    for famille_name, (fig_sub_bar, fig_sub_pie) in pareto_par_famille_figs.items():
        pptx_sections.append({
            "title": f"🔍 Pareto Niveau 2 — {famille_name}",
            "image": fig_to_png_bytes(fig_sub_bar),
            "image2": fig_to_png_bytes(fig_sub_pie),
        })

pptx_sections.append({
    "title": "🔁 Pareto par Fréquence des pannes",
    "caption": "Classement par NOMBRE de pannes (et non par durée) — équipements qui "
               "tombent en panne le plus souvent.",
    "image": fig_to_png_bytes(fig_freq),
    "table": freq_pareto,
})

if not paynter.empty:
    pptx_sections.append({
        "title": "📈 Paynter Chart — Évolution hebdomadaire des causes",
        "image": fig_to_png_bytes(fig3),
    })

if not alerts.empty:
    pptx_sections.append({
        "title": "🚨 Alertes machines critiques",
        "caption": f"Machine(s) dépassant le seuil de {threshold} min cumulées sur la période.",
        "table": alerts,
    })

pptx_sections.append({
    "title": "🗂️ Plan d'action pré-rempli",
    "caption": "Généré automatiquement à partir des équipements/causes les plus impactants — "
               "détail complet (cause racine, action, responsable, échéance) dans l'export Excel.",
    "table": action_plan_ppt,
    "table_header_color": action_plan_header_color,
})

buf_pptx = build_pptx_report(
    pptx_sections,
    title="Dashboard de diagnostic automatisé des temps d'arrêt",
    subtitle=f"Projet PFA — Génie Mécatronique | Rapport généré le {pd.Timestamp.today().date()}",
    closing_text="Développé dans le cadre d'un Projet de Fin d'Année — Génie Mécatronique. "
                 "Basé sur les données réelles de suivi downtime Versigent (Kaizen Event 4Q).",
)

st.download_button(
    "📦 Télécharger le rapport PowerPoint complet (prêt à projeter)",
    buf_pptx.getvalue(),
    file_name=f"Rapport_Downtime_Versigent_{pd.Timestamp.today().date()}.pptx",
    mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
    key="dl_all_pptx",
)

with st.expander("🗂️ Aperçu du plan d'action pré-rempli inclus dans la présentation"):
    st.dataframe(action_plan, use_container_width=True)
    st.caption(
        "Ces lignes sont générées automatiquement à partir des équipements/causes les plus "
        "impactants du Pareto. L'ingénieur maintenance n'a plus qu'à compléter la cause racine "
        "détaillée, l'action corrective, le responsable et l'échéance — au lieu de partir d'une "
        "page blanche."
    )

st.divider()
st.caption(
    "Développé dans le cadre d'un Projet de Fin d'Année — Génie Mécatronique. "
    "Basé sur les données réelles de suivi downtime Versigent (Kaizen Event 4Q)."
)
